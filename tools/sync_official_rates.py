import os
import sys
import sqlite3
import openpyxl

# Force UTF-8 stdout
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

DB_PATH = r"cusway.db"
EXCEL_PATH = r"관세청_품목번호별 관세율표_20260211.xlsx"

FTA_CODE_MAP = {
    'FUS': ('US', '한-미 FTA'),
    'FEU': ('IT', '한-EU FTA'),
    'FCN': ('CN', '한-중 FTA'),
    'FVN': ('VN', '한-베트남 FTA'),
    'FAU': ('AU', '한-호주 FTA'),
    'FCA': ('CA', '한-캐나다 FTA'),
    'FGB': ('GB', '한-영 FTA'),
    'FNZ': ('NZ', '한-뉴질랜드 FTA'),
    'FTR': ('TR', '한-튀르키예 FTA'),
    'FPE': ('PE', '한-페루 FTA'),
    'FCL': ('CL', '한-칠레 FTA'),
    'FIN': ('IN', '한-인도 CEPA'),
    'FSG': ('SG', '한-싱가포르 FTA'),
    'FAS': ('VN', '한-아세안 FTA'),
    'FRCJ': ('JP', 'RCEP (일본)'),
    'FRCAU': ('AU', 'RCEP (호주)'),
    'FRCNZ': ('NZ', 'RCEP (뉴질랜드)'),
    'FRCAS': ('VN', 'RCEP (아세안)'),
    'FID': ('ID', '한-인도네시아 CEPA'),
    'FKH': ('KH', '한-캄보디아 FTA'),
    'FPH': ('PH', '한-필리핀 FTA'),
    'FIL': ('IL', '한-이스라엘 FTA'),
    'FCO': ('CO', '한-콜롬비아 FTA'),
    'FCENI': ('NI', '한-중미 FTA (니카라과)'),
    'FCECR': ('CR', '한-중미 FTA (코스타리카)'),
    'FCEHN': ('HN', '한-중미 FTA (온두라스)'),
    'FCESV': ('SV', '한-중미 FTA (엘살바도르)'),
    'FCEPA': ('PA', '한-중미 FTA (파나마)')
}

def sync_compact_rates():
    print("=" * 80)
    print("       2026 관세청 공식 관세율표 고효율 최적화 동기화 엔진")
    print("=" * 80)
    
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    sheet_name = '2.12' if '2.12' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    
    items = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        raw_hs = str(row[0]).strip() if row[0] is not None else ""
        if not raw_hs: continue
        
        clean_hs = raw_hs.replace('.', '').replace('-', '')
        if len(clean_hs) < 10:
            clean_hs = clean_hs.zfill(10)
            
        rate_type = str(row[1]).strip() if row[1] is not None else ""
        try:
            rate_val = float(row[2]) if row[2] is not None else None
        except (ValueError, TypeError):
            rate_val = None
            
        try:
            specific_val = float(row[3]) if row[3] is not None else None
        except (ValueError, TypeError):
            specific_val = None
            
        unit_str = str(row[4]).strip() if (len(row) > 4 and row[4] is not None) else "kg"
        if not unit_str or unit_str in ['None', '1']: unit_str = "kg"
            
        if clean_hs not in items:
            items[clean_hs] = {
                'base': None,
                'wto': None,
                'base_sp': None,
                'wto_sp': None,
                'unit': unit_str,
                'ftas': {} # country_code -> (rate, fta_name, specific_val)
            }
            
        if rate_val is None and specific_val is None:
            continue
            
        if rate_type == 'A':
            items[clean_hs]['base'] = rate_val
            items[clean_hs]['base_sp'] = specific_val
        elif rate_type == 'C':
            items[clean_hs]['wto'] = rate_val
            items[clean_hs]['wto_sp'] = specific_val
        else:
            for prefix, (c_code, f_name) in FTA_CODE_MAP.items():
                if rate_type.startswith(prefix) and c_code in ['US', 'CN', 'VN', 'IT', 'JP', 'AU', 'CA', 'GB', 'IN', 'NZ', 'CL']:
                    # 만약 동일 국가에 복수 FTA가 있으면 더 유리한(낮은) 세율 우선 보존
                    if c_code in items[clean_hs]['ftas']:
                        existing_rate, _, _ = items[clean_hs]['ftas'][c_code]
                        if rate_val is not None and (existing_rate is None or rate_val < existing_rate):
                            items[clean_hs]['ftas'][c_code] = (rate_val, f_name, specific_val)
                    else:
                        items[clean_hs]['ftas'][c_code] = (rate_val, f_name, specific_val)
                    break

    records = []
    for hs_code, data in items.items():
        base_rate = data['base'] if data['base'] is not None else 8.0
        wto_rate = data['wto']
        base_sp = data['base_sp']
        wto_sp = data['wto_sp']
        unit = f"원/{data['unit']}"
        
        # 1. 기본 레코드 (BASE)
        rates_avail = [base_rate]
        if wto_rate is not None: rates_avail.append(wto_rate)
        rec_rate = min(rates_avail)
        
        # Determine duty formula & type for BASE/WTO
        duty_type = "AD_VALOREM"
        duty_formula = None
        sp_duty = wto_sp if wto_sp is not None else base_sp
        
        if wto_rate is not None and wto_sp is not None:
            duty_type = "ALTERNATIVE"
            duty_formula = f"{wto_rate}% 또는 {wto_sp:,.0f}{unit} 중 고액"
        elif base_rate is not None and base_sp is not None:
            duty_type = "ALTERNATIVE"
            duty_formula = f"{base_rate}% 또는 {base_sp:,.0f}{unit} 중 고액"
        elif sp_duty is not None and (base_rate is None and wto_rate is None):
            duty_type = "SPECIFIC"
            duty_formula = f"{sp_duty:,.0f}{unit}"
            
        records.append((hs_code, 'BASE', base_rate, wto_rate, None, '기본/WTO', rec_rate, sp_duty, unit if sp_duty else None, duty_type, duty_formula))
        
        # 2. 실제 유효한 FTA가 존재하는 국가들만 추가 적재
        for cntry, (fta_rate, fta_name, fta_sp) in data['ftas'].items():
            fta_avail = [base_rate]
            if wto_rate is not None: fta_avail.append(wto_rate)
            if fta_rate is not None: fta_avail.append(fta_rate)
            fta_rec_rate = min(fta_avail)
            
            cntry_duty_type = "AD_VALOREM"
            cntry_formula = None
            if fta_rate is not None and fta_sp is not None:
                cntry_duty_type = "ALTERNATIVE"
                cntry_formula = f"{fta_name}: {fta_rate}% 또는 {fta_sp:,.0f}{unit} 중 고액"
            elif duty_formula:
                cntry_duty_type = duty_type
                cntry_formula = duty_formula
                
            records.append((hs_code, cntry, base_rate, wto_rate, fta_rate, fta_name, fta_rec_rate, fta_sp or sp_duty, unit if (fta_sp or sp_duty) else None, cntry_duty_type, cntry_formula))

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Check if columns exist in hs_rate_master, if not recreate table
    cursor.execute("DROP TABLE IF EXISTS hs_rate_master")
    cursor.execute("""
        CREATE TABLE hs_rate_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hs_code TEXT NOT NULL,
            country_code TEXT NOT NULL,
            base_rate REAL,
            wto_rate REAL,
            fta_rate REAL,
            fta_name TEXT,
            recommended_rate REAL,
            specific_rate REAL,
            specific_unit TEXT,
            duty_type TEXT DEFAULT 'AD_VALOREM',
            duty_formula TEXT
        )
    """)
    
    print(f"💾 종가/종량 선택세율 포함 정밀 레코드 일괄 적재 중... (총 {len(records):,}개 세율 레코드)")
    cursor.executemany("""
        INSERT INTO hs_rate_master (hs_code, country_code, base_rate, wto_rate, fta_rate, fta_name, recommended_rate, specific_rate, specific_unit, duty_type, duty_formula)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, records)
    
    # 단일 복합 고속 인덱스 생성
    cursor.execute("CREATE INDEX IF NOT EXISTS ix_hs_rate_search ON hs_rate_master (hs_code, country_code);")
    
    conn.commit()
    conn.execute("VACUUM")
    conn.close()
    
    size_mb = os.path.getsize(DB_PATH) / (1024 * 1024)
    print(f"✨ 최적화 완료! 총 레코드: {len(records):,}건, DB 파일 크기: {size_mb:.2f} MB")

if __name__ == "__main__":
    sync_compact_rates()
